import requests
import os
import zipfile
import re
import glob
import csv
import sqlite3
import openpyxl
from openpyxl import load_workbook
import math

## Downloading the Medicare Hospital Compare Data and Loading it into SQL

staging_dir_name = "staging"
os.mkdir(staging_dir_name)

# Download the zip file from internet and then unzip it and save all the files into the staging directory
zip_url="https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip"
r = requests.get(zip_url)
zip_file_name = os.path.join(staging_dir_name, "Hospital_Revised_Flatfiles.zip")
zf = open(zip_file_name, "wb")
zf.write(r.content)
zf.close()
z = zipfile.ZipFile(zip_file_name,"r")
z.extractall(staging_dir_name)
z.close()

# Define a function to modify the table and column name as per the requirement
def Table_Column_naming(input_text,name_type):
    input_text = input_text.lower().replace(" ","_").replace("-","_").replace("%","pct").replace("/","_")
    if not re.match(r'^[a-z]',input_text):
        if name_type == "Table":
            input_text = "t_"+input_text
        elif name_type == "Column":
            input_text = "c_"+input_text
    return input_text

root_dir = os.getcwd()

# Fetch all the csv file name in the staging folder and save it into csv_list
os.chdir(os.path.join(root_dir,staging_dir_name))
csv_list = glob.glob('*.csv')

db_file_name = os.path.join(root_dir,"medicare_hospital_compare.db")

# Define a function to decode all the files using coding 'cp1252', then save to the csv file using coding 'utf-8'
def csv_encoding_transfer(csv_file_name):
    with open(csv_file_name, 'rt', encoding='cp1252') as f:
        input_data = f.read()
        if csv_file_name == 'FY2015_Percent_Change_in_Medicare_Payments.csv':
            input_data = input_data.replace('\0d"', '<=').replace('\n\0\n','\n').replace('"\0\t','",')
            input_data = '"' + input_data[2:].replace('\t\0 \0', '",')
        elif csv_file_name == 'MORT_READM_April2017.csv' or csv_file_name == 'PSI_April2017.csv':
            input_data = input_data.replace('\n ','\n"').replace('F,','F",')[0:-3]
    with open(csv_file_name, 'wt', encoding='utf-8') as r:
        for c in input_data:
            if c != '\0' and c != '\t':
                r.write(c)

# Define a function of processing the csv file to create table and save all the contents into the database file
def csv_file_processing(csv_file_name,db_file_name):
    csv_encoding_transfer(csv_file_name)
    table_name = Table_Column_naming(csv_file_name[0:-4],"Table")
    conn = sqlite3.connect(db_file_name)
    c1 = conn.cursor()
    with open(csv_file_name, 'r', encoding='utf-8') as f:
        #reader = csv.reader(f,delimiter=',', quoting=csv.QUOTE_NONE)
        reader = csv.reader(f)
        i = 1
        for row in reader:
            if i == 1:
                column_name_list = [Table_Column_naming(x,"Column") for x in row]
                sql_str = 'create table if not exists '+table_name+' (\n'+' text,\n'.join(column_name_list)+' text\n)'
                c1.execute(sql_str)
                i+=1
            else:
                #sql_str = 'insert into '+table_name+' '+str(tuple(column_name_list))+' values ('+('?,'*len(row))[0:-1]+')'
                sql_str = 'insert into '+table_name+' values ('+('?,'*len(row))[0:-1]+')'
                sql_tuple = tuple(row)
                c1.execute(sql_str,sql_tuple)
        conn.commit()

for csv_file_name in csv_list:
    csv_file_processing(csv_file_name,db_file_name)

## Download MS Excel Workbook of In House Proprietary Hospital Rankings and Focus List of States

# Define a function of processing excel file to create and save contents into database file
def excel_sheet_processing(sheet_name,db_file_name):
    table_name = sheet_name
    conn = sqlite3.connect(db_file_name)
    c1 = conn.cursor()
    sheet = wb.get_sheet_by_name(sheet_name)
    i = 1
    while sheet.cell(row = i, column = 1).value != None:
        if i == 1:
            if table_name == "Hospital National Ranking":
                sql_str = 'create table if not exists "'+table_name+ \
                '" (\n"'+sheet.cell(row = i, column = 1).value+'" text,\n"'+ \
                sheet.cell(row = i, column = 2).value+'" integer\n)'
            else:
                sql_str = 'create table if not exists "'+table_name+ \
                '" (\n"'+sheet.cell(row = i, column = 1).value+'" text,\n"'+ \
                sheet.cell(row = i, column = 2).value+'" text\n)'
            c1.execute(sql_str)
        else:
            sql_str = 'insert into "'+table_name+'" values ("'+ \
            str(sheet.cell(row = i, column = 1).value)+'","'+str(sheet.cell(row = i, column = 2).value)+'")'
            c1.execute(sql_str)
        i+=1
    conn.commit()

# Download excel file from internet and save it to the local directory
Excel_url="http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"
r = requests.get(Excel_url)
xf = open("hospital_ranking_focus_states.xlsx","wb")
xf.write(r.content)
xf.close()

wb = openpyxl.load_workbook("hospital_ranking_focus_states.xlsx")
for sheet_name in wb.get_sheet_names():
    excel_sheet_processing(sheet_name,db_file_name)

## Create the Hospital Ranking MS Excel Workbook

# Define a function to wirte the database inquiry result into "hospital_ranking" excel sheet
def write2sheet(rows,sheet):
    sheet.cell(row = 1, column = 1, value = 'Provider ID')
    sheet.cell(row = 1, column = 2, value = 'Hospital Name')
    sheet.cell(row = 1, column = 3, value = 'City')
    sheet.cell(row = 1, column = 4, value = 'State')
    sheet.cell(row = 1, column = 5, value = 'County')
    row_num = 2
    for row in rows:
        for col_num in range(len(row)):
            if row_num < 102:
                sheet.cell(row = row_num, column = col_num+1, value = row[col_num])
        row_num += 1

# Perform SQL inquiry to get contents of first sheet "Nationwide"
conn = sqlite3.connect(db_file_name)
c1 = conn.cursor()
sql_str = 'select "Provider ID",hospital_name,city,state,county_name from hospital_general_information ' + \
            'join "Hospital National Ranking" ' + \
            'on hospital_general_information.provider_id = "Hospital National Ranking"."Provider ID" ' + \
            'where "Hospital National Ranking".Ranking < 101 ' + \
            'order by "Hospital National Ranking".Ranking'
c1.execute(sql_str)
rows = c1.fetchall()

# Create and write the SQL inquriy result into the first sheet "Nationwide"
new_wb = openpyxl.Workbook()
sheet_1 = new_wb.create_sheet("Nationwide")
write2sheet(rows,sheet_1)
new_wb.remove_sheet(new_wb.get_sheet_by_name("Sheet"))
new_wb.save(os.path.join(root_dir,"hospital_ranking.xlsx"))

# Fetch the focus state list and sort it in alphabetic order
sql_str = 'select "State Name" from "Focus States"'
c1.execute(sql_str)
rows = c1.fetchall()
Focus_state = [i[0] for i in rows]
Focus_state.sort()

# Create and write the SQL inquriy result into the sheet per focus state
for state in Focus_state:
    new_wb = load_workbook(os.path.join(root_dir,"hospital_ranking.xlsx"))
    sheet_name = new_wb.create_sheet(state)
    sql_str = 'select "Provider ID",hospital_name,city,state,county_name from hospital_general_information ' + \
                'join "Hospital National Ranking" ' + \
                'on hospital_general_information.provider_id = "Hospital National Ranking"."Provider ID" ' + \
                'join "Focus States" ' + \
                'on "Focus States"."State Abbreviation" = hospital_general_information.state ' + \
                'where "Focus States"."State name" = "' + state + '" ' + \
                'order by "Hospital National Ranking".Ranking'
    c1.execute(sql_str)
    rows = c1.fetchall()
    write2sheet(rows,sheet_name)
    new_wb.save(os.path.join(root_dir,"hospital_ranking.xlsx"))

## Create the Measures Statistical Analysis MS Excel Workbook

# Query data from timely_and_effective_care___hospital table and remove those non-numeric scores
sql_str = 'select hospital_name,state,measure_id,measure_name,score from timely_and_effective_care___hospital'
c1.execute(sql_str)
rows = c1.fetchall()
new_rows = [row for row in rows if len(row[4]) < 5]

# Save the query result into a new table in database, named as filtered_timely_and_effective_care_hospital
i = 1
for row in new_rows:
    if i == 1:
        sql_str = 'create table if not exists filtered_timely_and_effective_care_hospital (\n' + \
                    'hospital_name text,\n' + \
                    'state text,\n' + \
                    'measure_id text,\n' + \
                    'measure_name text,\n' + \
                    'score integer\n)'
        c1.execute(sql_str)
        i+=1
    else:
        sql_str = 'insert into filtered_timely_and_effective_care_hospital values ('+('?,'*len(row))[0:-1]+')'
        sql_tuple = tuple(row)
        c1.execute(sql_str,sql_tuple)
conn.commit()

# Define a function to wirte the database inquiry result into "measure_statistics" excel sheet
def stats_write2sheet(rows,sheet):
    sheet.cell(row = 1, column = 1, value = 'Measure ID')
    sheet.cell(row = 1, column = 2, value = 'Measure Name')
    sheet.cell(row = 1, column = 3, value = 'Minimum')
    sheet.cell(row = 1, column = 4, value = 'Maximum')
    sheet.cell(row = 1, column = 5, value = 'Average')
    sheet.cell(row = 1, column = 6, value = 'Standard Deviation')
    row_num = 2
    for row in rows:
        for col_num in range(len(row)):
            sheet.cell(row = row_num, column = col_num+1, value = row[col_num])
        row_num+=1


# Define a function to perform SQL inquiry to calculate standard deviation
def stats_SQL(db_file_name, data_range):
    conn = sqlite3.connect(db_file_name)
    c1 = conn.cursor()
    if data_range == "Nationwide":
        sql_str = 'select measure_id,measure_name,MIN(score),MAX(score),AVG(score)' + \
                  'from filtered_timely_and_effective_care_hospital ' + \
                  'group by measure_id ' + \
                  'order by measure_name'
    else:
        sql_str = 'select measure_id,measure_name,MIN(score),MAX(score),AVG(score)' + \
                  'from filtered_timely_and_effective_care_hospital ' + \
                  'join "Focus States" ' + \
                  'on "Focus States"."State Abbreviation" = filtered_timely_and_effective_care_hospital.state ' + \
                  'where "Focus States"."State Name" = "' + data_range + \
                  '" group by measure_id ' + \
                  'order by measure_name'
    c1.execute(sql_str)
    rows = c1.fetchall()

    new_rows = []
    for row in rows:
        measure_id = row[0]
        AVG = row[4]
        if data_range == "Nationwide":
            sql_str = 'select measure_id,score from filtered_timely_and_effective_care_hospital ' + \
                      'where measure_id = "' + measure_id + '"'
        else:
            sql_str = 'select measure_id,score from filtered_timely_and_effective_care_hospital ' + \
                      'join "Focus States" ' + \
                      'on "Focus States"."State Abbreviation" = filtered_timely_and_effective_care_hospital.state ' + \
                      'where "Focus States"."State Name" = "' + data_range + \
                      '" and measure_id = "' + measure_id + '"'
        c1.execute(sql_str)
        sub_rows = c1.fetchall()
        SUM = 0
        COUNT = 0
        for sub_row in sub_rows:
            SUM += pow((sub_row[1] - AVG), 2)
            COUNT += 1
        new_rows.append(row + (math.sqrt(SUM / COUNT),))
    return new_rows

# Create MS Excel Workbook named as “measure_statistics.xlsx” and it's first sheet "Nationwide"
new_wb = openpyxl.Workbook()
stats_sheet1 = new_wb.create_sheet("Nationwide")
stats_write2sheet(stats_SQL(db_file_name,"Nationwide"),stats_sheet1)
new_wb.remove_sheet(new_wb.get_sheet_by_name("Sheet"))
new_wb.save(os.path.join(root_dir,"measure_statistics.xlsx"))

# Write the focus state sheet into the MS Excel Workbook “measure_statistics.xlsx”
for state in Focus_state:
    new_wb = load_workbook(os.path.join(root_dir,"measure_statistics.xlsx"))
    sheet_name = new_wb.create_sheet(state)
    stats_write2sheet(stats_SQL(db_file_name,state),sheet_name)
    new_wb.save(os.path.join(root_dir,"measure_statistics.xlsx"))
